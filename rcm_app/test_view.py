from django.shortcuts import render, redirect

# Create your views here.
# rcm_app/views.py
from django.http import HttpResponse

def home(request):
    return render(request, 'home.html')

from django.contrib.auth import login
from django.contrib.auth.views import LoginView, LogoutView
from .forms import RegisterForm

def register_view(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)  # Log user in after registration
            return redirect("home")
    else:
        form = RegisterForm()
    return render(request, "register.html", {"form": form})


import pandas as pd
from django.db import connection
import re
from datetime import datetime
from django.shortcuts import render
from .forms import ExcelUploadForm


def sanitize_column_name(col_name):
    """Sanitize column name to avoid issues with SQL syntax."""
    return re.sub(r'\W|^(?=\d)', '_', col_name)


def convert_to_sql_compatible(value):
    """Convert values to SQL-compatible formats."""
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return value
    return str(value)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import ExcelUploadForm
from .models import ExcelUpload, ExcelData
import pandas as pd
import numpy as np
from datetime import datetime


def convert_to_serializable(value):
    """Convert pandas and numpy types to Python native types"""
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if isinstance(value, (np.integer)):
        return int(value)
    if isinstance(value, (np.floating)):
        return float(value)
    if isinstance(value, (np.ndarray)):
        return value.tolist()
    return value


#@login_required
def upload_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file = request.FILES['file']

                # Read Excel file
                df = pd.read_excel(file, engine='openpyxl')

                # Convert all values to JSON-serializable format
                serializable_data = []
                for _, row in df.iterrows():
                    serializable_row = {col: convert_to_serializable(row[col]) for col in df.columns}
                    serializable_data.append(serializable_row)

                # Create ExcelUpload object with current user
                upload = ExcelUpload.objects.create(
                    user=request.user,  # Assign user
                    file_name=file.name,
                    row_count=len(df),
                    columns={col: str(df[col].dtype) for col in df.columns}
                )

                # Bulk insert ExcelData
                ExcelData.objects.bulk_create([
                    ExcelData(upload=upload, data=row_data)
                    for row_data in serializable_data
                ])

                return render(request, 'upload_success.html', {
                    'upload': upload,
                    'columns': df.columns.tolist(),
                    'row_count': len(df)
                })

            except Exception as e:
                error_msg = str(e)
                if 'No such file or directory' in error_msg:
                    error_msg = "Please select a file to upload."
                return render(request, 'upload.html', {
                    'form': form,
                    'error': f'Error processing file: {error_msg}'
                })
    else:
        form = ExcelUploadForm()

    return render(request, 'upload.html', {'form': form})

from django.shortcuts import render
from .models import ExcelData


def classify_payment_status(row):
    balance = float(row.get('Balance amt', 0) or 0)
    charge = float(row.get('Charge amt', 0) or 0)
    payments = float(row.get('Payments', 0) or 0)
    status = row.get('Status', '').lower()

    if balance < 0:
        return "Negative balance"
    elif balance == 0 and charge == 0 and status in ['canceled', 'closed']:
        return "Canceled Trip"
    elif balance == 0 and charge > 0 and payments > 0:
        return "Paid & Closed"
    elif balance == 0 and payments == 0:
        return "Adjusted"
    elif payments > 0:
        return "Partially paid"
    elif charge != 0 or balance != 0:
        return "Unpaid"
    return ""


def classify_ar_status(row, payment_status):
    status = row.get('Status', '').lower()
    payor = row.get('Current Payor category', '').lower()
    pri_payor = row.get('Primary Payor Category', '').lower()
    schedule_track = row.get('schedule/Track', '').lower()

    if payment_status == "Negative balance":
        return "Negative Ins AR"
    elif (
        (float(row.get('Balance amt', 0) or 0) == 0 and float(row.get('Charge amt', 0) or 0) == 0 and status in ['canceled', 'closed']) or
        payment_status == "Canceled Trip"
    ):
        return "Canceled Trip"
    elif payment_status == "Paid & Closed" and pri_payor == 'patient' and payor == 'patient':
        return "Closed - Pt Pri"
    elif payment_status == "Paid & Closed":
        return "Closed - Ins Pri"
    elif payment_status == "Adjusted":
        return "Adjusted & Closed"
    elif payor == "patient" and "denials" not in schedule_track and "waystar" not in schedule_track:
        return "Open - Pt AR"
    return "Open - Ins AR"


def classify_claim_status(row, payment_status, ar_status):
    """
    Determine claim status based on the 12-step logic provided
    Returns tuple: (claim_status, debug_steps)
    """
    debug_steps = []

    # Extract relevant fields with defaults
    balance = float(row.get('Balance Due', 0) or 0)
    charge = float(row.get('Net Charges', 0) or 0)
    payments = float(row.get('Payments', 0) or 0)  # Fixed this line
    status = row.get('Status', '').lower()
    payor = row.get('Current Payor Category', '').lower()
    pri_payor = row.get('Primary Payor Category', '').lower()
    schedule_track = row.get('Schedule/Track', '').lower()

    # Rest of the function remains the same...
    # Step 1: Negative balance
    if balance < 0:
        debug_steps.append("Step 1: Negative balance detected")
        return "Negative balance", debug_steps

    # Step 2: Canceled but closed
    if (payment_status == "Canceled Trip" and status == 'closed') or \
            (balance == 0 and charge == 0 and status == 'closed'):
        debug_steps.append("Step 2: Canceled with closed status")
        return "Canceled but Status Closed", debug_steps

    # Step 3: Canceled with Posting
    if payment_status == "Canceled Trip" and \
            (balance != 0 or charge != 0 or payments != 0):
        debug_steps.append("Step 3: Canceled with financial activity")
        return "Canceled with Posting", debug_steps

    # Step 4: Canceled Trip
    if balance == 0 and charge == 0 and status == 'canceled':
        debug_steps.append("Step 4: Canceled trip with no activity")
        return "Canceled Trip", debug_steps

    # Step 5: New Trips
    if status == 'new' or 'emsmart processed' in schedule_track:
        debug_steps.append("Step 5: New trip detected")
        return "New Trips", debug_steps

    # Step 6: Paid & Closed
    if payment_status == "Paid & Closed" or \
            (balance == 0 and charge > 0 and payments > 0):
        debug_steps.append("Step 6: Paid and closed claim")
        return "Paid & Closed", debug_steps

    # Step 7: Adjusted & Closed
    if payment_status == "Adjusted":
        debug_steps.append("Step 7: Adjusted claim")
        return "Adjusted & Closed", debug_steps

    # Step 8: Patient Signature Requested npp signature required
    if ar_status == "Open - Pt AR" and \
            ('signature required' in schedule_track or 'npp' in schedule_track):
        debug_steps.append("Step 8: Patient signature required")
        return "Pt Sign requested", debug_steps

    # Step 9: Billed to Patient - Primary
    if (payor == 'patient' and pri_payor == 'patient' and \
        not any(x in schedule_track for x in ['waystar', 'denials', 'automatic crossover'])) or \
            (ar_status == "Open - Pt AR" and pri_payor == 'patient'):
        debug_steps.append("Step 9: Billed to patient (primary)")
        return "Billed to Pt - Pri", debug_steps

    # Step 10: Billed to Patient - Secondary
    if ar_status == "Open - Pt AR" and \
            payor == 'patient' and pri_payor != 'patient':
        debug_steps.append("Step 10: Billed to patient (secondary)")
        return "Billed to Pt - Sec", debug_steps

    # Step 11: Billed to Insurance - Primary
    if ar_status == "Open - Ins AR" and \
            payor == pri_payor and \
            'automatic crossover' not in schedule_track:
        debug_steps.append("Step 11: Billed to insurance (primary)")
        return "Billed to Ins - Pri", debug_steps

    # Step 12: Billed to Insurance - Secondary
    if ar_status == "Open - Ins AR":
        debug_steps.append("Step 12: Billed to insurance (secondary)")
        return "Billed to Ins - Sec", debug_steps

    debug_steps.append("No matching claim status found - defaulting")
    return "Unclassified", debug_steps

def test_display_data_verbose(request):
    # Fetch first 50 rows
    queryset = ExcelData.objects.filter(upload__user=request.user)[:50]

    processed_data = []

    print("\n======= STARTING CLASSIFICATION PROCESS =======")

    for i, row in enumerate(queryset, 1):
        d = row.data

        # Extract key fields (with defaults to avoid errors)
        balance = float(d.get('Balance Due', 0) or 0)
        charge = float(d.get('Net Charges', 0) or 0)
        payments = float(d.get('Payments', 0) or 0)
        status = d.get('Status', '').lower()
        payor = d.get('Current Payor Category', '').lower()
        pri_payor = d.get('Primary Payor Category', '').lower()
        schedule_track = d.get('Schedule/Track', '').lower()

        print(f"\n--- Row {i} ---")
        print(f"Key Values: Balance={balance}, Charge={charge}, Payments={payments}, Status='{status}'")
        print(f"Payor Info: Current='{payor}', Primary='{pri_payor}', Schedule='{schedule_track}'")

        # Payment Status Classification
        print("\n1. Determining Payment Status:")
        if balance < 0:
            ps = "Negative balance"
            print(f"  - Rule: Balance ({balance}) < 0 → '{ps}'")
        elif balance == 0 and charge == 0 and status in ['canceled', 'closed']:
            ps = "Canceled Trip"
            print(f"  - Rule: Zero balance & charge + status '{status}' → '{ps}'")
        elif balance == 0 and charge > 0 and payments > 0:
            ps = "Paid & Closed"
            print(f"  - Rule: Zero balance with payments → '{ps}'")
        elif balance == 0 and payments == 0:
            ps = "Adjusted"
            print(f"  - Rule: Zero balance without payments → '{ps}'")
        elif payments > 0:
            ps = "Partially paid"
            print(f"  - Rule: Payments exist but balance remains → '{ps}'")
        else:
            ps = "Unpaid"
            print(f"  - Rule: Default case → '{ps}'")

        # AR Status Classification
        print("\n2. Determining AR Status:")
        if ps == "Negative balance":
            ars = "Negative Ins AR"
            print(f"  - Rule: Payment Status is '{ps}' → '{ars}'")
        elif (balance == 0 and charge == 0 and status in ['canceled', 'closed']) or ps == "Canceled Trip":
            ars = "Canceled Trip"
            print(f"  - Rule: Canceled trip conditions → '{ars}'")
        elif ps == "Paid & Closed" and pri_payor == 'patient' and payor == 'patient':
            ars = "Closed - Pt Pri"
            print(f"  - Rule: Paid & patient primary → '{ars}'")
        elif ps == "Paid & Closed":
            ars = "Closed - Ins Pri"
            print(f"  - Rule: Paid & non-patient primary → '{ars}'")
        elif ps == "Adjusted":
            ars = "Adjusted & Closed"
            print(f"  - Rule: Payment Status is '{ps}' → '{ars}'")
        elif payor == "patient" and "denials" not in schedule_track and "waystar" not in schedule_track:
            ars = "Open - Pt AR"
            print(f"  - Rule: Patient payor without denials → '{ars}'")
        else:
            ars = "Open - Ins AR"
            print(f"  - Rule: Default case → '{ars}'")

        # Claim Status Classification
        print("\n3. Determining Claim Status:")
        cs, cs_debug = classify_claim_status(d, ps, ars)
        for step in cs_debug:
            print(f"  - {step}")
        print(f"  - Final Claim Status: '{cs}'")

        print(f"\nFinal Classification: Payment='{ps}', AR='{ars}', Claim='{cs}'")

        processed_data.append({
            **d,
            'Payment Status': ps,
            'AR Status': ars,
            'Claim Status': cs
        })

    print("\n======= CLASSIFICATION COMPLETE =======")
    return render(request, 'testing.html', {'data': processed_data})


import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def download_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExcelData Preview"

    data = [
        {
            **row.data,
            'Payment Status': (ps := classify_payment_status(row.data)),
            'AR Status': classify_ar_status(row.data, ps),
        }
        for row in ExcelData.objects.filter(upload__user=request.user)[:50]
    ]

    if not data:
        return HttpResponse("No data to export.", status=400)

    # Write headers
    headers = list(data[0].keys())
    ws.append(headers)

    # Write rows
    for row in data:
        ws.append([row.get(col, '') for col in headers])

    # Adjust column widths
    for col_num, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=excel_data_preview.xlsx'
    wb.save(response)
    return response
from django.template.loader import get_template
from xhtml2pdf import pisa

def download_pdf(request):
    data = [
        {
            **row.data,
            'Payment Status': (ps := classify_payment_status(row.data)),
            'AR Status': classify_ar_status(row.data, ps),
        }
        for row in ExcelData.objects.filter(upload__user=request.user)[:50]
    ]

    template = get_template('pdf_template.html')  # We'll create this
    html = template.render({'data': data})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="excel_data_preview.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response

from django.shortcuts import render
from collections import Counter, defaultdict
from datetime import datetime

def dashboard_view(request):
    rows = ExcelData.objects.filter(upload__user=request.user)


    processed = []
    net_charges = 0
    total_payments = 0
    total_balance = 0
    ar_days = []

    for row in rows:
        d = row.data
        ps = classify_payment_status(d)
        ars = classify_ar_status(d, ps)

        net = float(d.get('Net Charges', 0))
        paid = float(d.get('Payments', 0))
        balance = float(d.get('Balance Due', 0))

        net_charges += net
        total_payments += paid
        total_balance += balance

        # Aging Days
        aging_str = d.get('Aging Date')
        if aging_str:
            try:
                aging_date = datetime.strptime(aging_str, "%Y-%m-%d")
                ar_days.append((datetime.now() - aging_date).days)
            except Exception:
                pass

        d['Payment Status'] = ps
        d['AR Status'] = ars
        processed.append(d)

    payment_counts = Counter(d['Payment Status'] for d in processed)
    ar_counts = Counter(d['AR Status'] for d in processed)
    payor_counts = Counter(d.get('Pri Payor Category', 'Unknown') for d in processed)
    track_counts = Counter(d.get('Schedule/Track', 'Unknown') for d in processed)

    import_trends = defaultdict(lambda: {'charges': 0, 'payments': 0})
    for d in processed:
        ym = d.get('Import Date YM')
        import_trends[ym]['charges'] += float(d.get('Net Charges', 0))
        import_trends[ym]['payments'] += float(d.get('Payments', 0))

    avg_ar_days = round(sum(ar_days) / len(ar_days), 1) if ar_days else 0

    return render(request, 'dashboard.html', {
        'total_claims': len(processed),
        'net_charges': net_charges,
        'total_payments': total_payments,
        'total_balance': total_balance,
        'avg_ar_days': avg_ar_days,
        'payment_counts': dict(payment_counts),
        'ar_counts': dict(ar_counts),
        'payor_counts': dict(payor_counts),
        'track_counts': dict(track_counts),
        'import_trends': dict(import_trends),
    })
